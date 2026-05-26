# view/admin/export_view.py
"""
ExportView — Export clinic data to Excel (.xlsx)

Exports available:
  1. Patients list
  2. Appointments (today / this month / all)
  3. Prescriptions
  4. Revenue summary

Each export opens a Save dialog then opens the file automatically.
Uses openpyxl for rich formatting (headers, colours, column widths).
"""

import customtkinter as ctk
from tkinter import filedialog, messagebox
import os
from datetime import datetime
from view.sidebar import Sidebar, Toast

# openpyxl — installed via: pip install openpyxl
try:
    from openpyxl import Workbook                                   # type: ignore
    from openpyxl.styles import Border, Side                        # type: ignore
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

BG             = "#F7F8FA"
PANEL_BG       = "#FFFFFF"
PANEL_ALT      = "#F1F5F9"
BORDER         = "#E2E6ED"
TEXT_PRIMARY   = "#1A202C"
TEXT_SECONDARY = "#64748B"
ACCENT         = "#2563EB"
ACCENT_LIGHT   = "#DBEAFE"
ACCENT_HOVER   = "#1D4ED8"
SUCCESS        = "#16A34A"
SUCCESS_LIGHT  = "#DCFCE7"
WARNING        = "#D97706"
WARNING_LIGHT  = "#FEF3C7"
DANGER         = "#DC2626"
PURPLE         = "#7C3AED"
PURPLE_LIGHT   = "#EDE9FE"

FONT = "Helvetica"


# ─────────────────────────────────────────────────────────────────────────────
class ExportView(ctk.CTkFrame):

    def __init__(self, parent, controller):
        super().__init__(parent, fg_color=BG)
        self.grid(row=0, column=0, sticky="nsew")
        parent.grid_rowconfigure(0, weight=1)
        parent.grid_columnconfigure(0, weight=1)

        self.controller = controller

        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=0)
        self.grid_columnconfigure(1, weight=1)

        self._sidebar = Sidebar(self, controller, active="export")
        self._sidebar.grid(row=0, column=0, sticky="ns")

        self._content = ctk.CTkFrame(self, fg_color=BG)
        self._content.grid(row=0, column=1, sticky="nsew")
        self._content.grid_rowconfigure(1, weight=1)
        self._content.grid_columnconfigure(0, weight=1)

        self._build_topbar()
        self._build_body()

    # ── Top bar ───────────────────────────────────────────────────────────

    def _build_topbar(self):
        bar = ctk.CTkFrame(self._content, fg_color=PANEL_BG, height=64,
                           border_width=1, border_color=BORDER, corner_radius=0)
        bar.grid(row=0, column=0, sticky="ew")
        bar.grid_propagate(False)
        bar.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(bar, text="📤  Export Data",
                     font=ctk.CTkFont(FONT, 20, "bold"),
                     text_color=TEXT_PRIMARY
                     ).grid(row=0, column=0, padx=24, sticky="w")

        ctk.CTkLabel(bar,
                     text="Export clinic data to Excel (.xlsx) for reporting and analysis",
                     font=ctk.CTkFont(FONT, 12),
                     text_color=TEXT_SECONDARY
                     ).grid(row=0, column=1, padx=8, sticky="w")

    # ── Body ──────────────────────────────────────────────────────────────

    def _build_body(self):
        body = ctk.CTkScrollableFrame(self._content, fg_color=BG)
        body.grid(row=1, column=0, sticky="nsew", padx=24, pady=20)
        body.grid_columnconfigure(0, weight=1)
        body.grid_columnconfigure(1, weight=1)

        EXPORTS = [
            {
                "title":    "👤  Patients",
                "desc":     "All registered patients with their personal details, blood type, wilaya, and contact info.",
                "fields":   "ID · Full Name · DOB · Age · Gender · Blood · Phone · Email · Wilaya · Address · Allergies · Registered",
                "color":    ACCENT,
                "bg":       ACCENT_LIGHT,
                "btn_text": "Export Patients",
                "action":   self._export_patients,
                "row": 0, "col": 0,
            },
            {
                "title":    "📅  Appointments",
                "desc":     "Appointment records with patient names, visit types, status, diagnosis and billing totals.",
                "fields":   "ID · Patient · Date · Time · Visit Type · Status · Complaint · Diagnosis · Total DA",
                "color":    PURPLE,
                "bg":       PURPLE_LIGHT,
                "btn_text": "Export Appointments",
                "action":   self._export_appointments,
                "row": 0, "col": 1,
            },
            {
                "title":    "📋  Prescriptions",
                "desc":     "All prescriptions with medicines, dosages, frequencies and durations per patient.",
                "fields":   "Rx ID · Patient · Date · Medicine · Dosage · Frequency · Duration · Instructions",
                "color":    SUCCESS,
                "bg":       SUCCESS_LIGHT,
                "btn_text": "Export Prescriptions",
                "action":   self._export_prescriptions,
                "row": 1, "col": 0,
            },
            {
                "title":    "💰  Revenue Report",
                "desc":     "Daily and monthly revenue summary from completed appointments with service breakdown.",
                "fields":   "Date · Appointments · Completed · Revenue DA · Avg per Visit",
                "color":    WARNING,
                "bg":       WARNING_LIGHT,
                "btn_text": "Export Revenue",
                "action":   self._export_revenue,
                "row": 1, "col": 1,
            },
            {
                "title":    "🩺  Services Catalog",
                "desc":     "Full services catalog with codes, categories, prices and durations.",
                "fields":   "Code · Name · Category · Price DA · Duration (min) · Description",
                "color":    "#0D9488",
                "bg":       "#CCFBF1",
                "btn_text": "Export Services",
                "action":   self._export_services,
                "row": 2, "col": 0,
            },
            {
                "title":    "📊  Full Report",
                "desc":     "Complete clinic report: all patients, appointments, prescriptions and revenue in one workbook with multiple sheets.",
                "fields":   "All sheets combined — ready for management reporting",
                "color":    DANGER,
                "bg":       "#FEE2E2",
                "btn_text": "Export Full Report",
                "action":   self._export_full_report,
                "row": 2, "col": 1,
            },
        ]

        for cfg in EXPORTS:
            self._build_export_card(body, cfg)


    def _check_openpyxl(self) -> bool:
        if not _OPENPYXL_OK:
            from tkinter import messagebox
            messagebox.showerror(
                "Missing Library",
                "openpyxl is not installed.\n"
                "Run:  pip install openpyxl")
            return False
        return True

    def _build_export_card(self, parent, cfg):
        card = ctk.CTkFrame(parent, fg_color=PANEL_BG,
                            border_width=1, border_color=BORDER,
                            corner_radius=12)
        card.grid(row=cfg["row"], column=cfg["col"],
                  sticky="nsew", padx=8, pady=8)
        card.grid_columnconfigure(0, weight=1)

        # Colour header strip
        hdr = ctk.CTkFrame(card, fg_color=cfg["bg"],
                           corner_radius=0, height=52)
        hdr.grid(row=0, column=0, sticky="ew")
        hdr.grid_propagate(False)
        hdr.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(hdr, text=cfg["title"],
                     font=ctk.CTkFont(FONT, 15, "bold"),
                     text_color=cfg["color"], anchor="w"
                     ).grid(row=0, column=0, padx=16, pady=14, sticky="w")

        # Description
        ctk.CTkLabel(card, text=cfg["desc"],
                     font=ctk.CTkFont(FONT, 12),
                     text_color=TEXT_PRIMARY,
                     anchor="w", wraplength=340, justify="left"
                     ).grid(row=1, column=0, padx=16, pady=(14, 6), sticky="w")

        # Fields preview
        ctk.CTkLabel(card,
                     text=f"Columns: {cfg['fields']}",
                     font=ctk.CTkFont(FONT, 10),
                     text_color=TEXT_SECONDARY,
                     anchor="w", wraplength=340, justify="left"
                     ).grid(row=2, column=0, padx=16, pady=(0, 14), sticky="w")

        # Export button
        ctk.CTkButton(card, text=f"📥  {cfg['btn_text']}",
                      font=ctk.CTkFont(FONT, 13, "bold"),
                      fg_color=cfg["color"],
                      hover_color=cfg["color"],
                      height=40, corner_radius=8,
                      command=cfg["action"]
                      ).grid(row=3, column=0, sticky="ew",
                             padx=16, pady=(0, 16))

    # ── Export helpers ────────────────────────────────────────────────────

    def _ask_save(self, default_name: str) -> str | None:
        """Open save dialog and return path, or None if cancelled."""
        return filedialog.asksaveasfilename(
            title="Save Excel File",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel Workbook", "*.xlsx"),
                       ("All Files", "*.*")],
        )

    def _open_file(self, path: str):
        """Open the saved file with the default application."""
        try:
            os.startfile(path)
        except AttributeError:
            try:
                import subprocess
                subprocess.Popen(["xdg-open", path])
            except Exception:
                pass

    def _make_wb(self):
        """Create a new openpyxl workbook."""
        return Workbook()

    def _style_header_row(self, ws, row: int, col_count: int,
                           bg_hex: str = "2563EB"):
        """Apply bold white text on coloured background to header row."""

        fill  = PatternFill("solid", fgColor=bg_hex.lstrip("#"))
        font  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        align = Alignment(horizontal="center", vertical="center")
        thin  = Side(style="thin", color="FFFFFF")
        bdr   = Border(bottom=thin)

        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font   = font
            cell.fill   = fill
            cell.alignment = align
            cell.border = bdr

    def _style_data_rows(self, ws, start_row: int, end_row: int,
                          col_count: int):
        """Alternate row colours for data rows."""

        fills = [
            PatternFill("solid", fgColor="FFFFFF"),
            PatternFill("solid", fgColor="F1F5F9"),
        ]
        font  = Font(name="Arial", size=10)
        align = Alignment(vertical="center")

        for row in range(start_row, end_row + 1):
            fill = fills[(row - start_row) % 2]
            for col in range(1, col_count + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill      = fill
                cell.font      = font
                cell.alignment = align

    def _add_clinic_header(self, ws, title: str, col_span: int):
        """Add clinic name + export title at the top of the sheet."""

        try:
            clinic = self.controller.get_app_name()
        except Exception:
            clinic = "Clinic Manager"

        # Row 1 — clinic name
        ws.merge_cells(f"A1:{get_column_letter(col_span)}1")
        c1 = ws["A1"]
        c1.value     = clinic.upper()
        c1.font      = Font(bold=True, size=14, name="Arial", color="FFFFFF")
        c1.fill      = PatternFill("solid", fgColor="0D9488")
        c1.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # Row 2 — export title + date
        ws.merge_cells(f"A2:{get_column_letter(col_span)}2")
        c2 = ws["A2"]
        c2.value     = f"{title}   —   Exported {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        c2.font      = Font(size=10, name="Arial", color="64748B", italic=True)
        c2.fill      = PatternFill("solid", fgColor="F0FDF4")
        c2.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20

    # ══════════════════════════════════════════════════════════════════════
    # EXPORT: PATIENTS
    # ══════════════════════════════════════════════════════════════════════

    def _export_patients(self):
        if not self._check_openpyxl(): return
        if not self._check_openpyxl(): return
        if not self._check_openpyxl(): return
        if not self._check_openpyxl(): return
        if not self._check_openpyxl(): return
        if not self._check_openpyxl(): return
        path = self._ask_save(
            f"patients_{datetime.now().strftime('%Y%m%d')}.xlsx")
        if not path:
            return

        try:
            rows = self.controller.db.fetch_all(
                """SELECT patient_id, full_name, date_of_birth, gender,
                          phone, email, wilaya, blood_type,
                          allergies, address, created_at
                   FROM patients ORDER BY full_name"""
            ) or []
        except Exception as e:
            messagebox.showerror("Export Error", str(e)); return

        from datetime        import date

        wb = self._make_wb()
        ws = wb.active
        ws.title = "Patients"

        COLS = ["ID", "Full Name", "Date of Birth", "Age", "Gender",
                "Blood Type", "Phone", "Email", "Wilaya",
                "Address", "Allergies", "Registered"]
        N = len(COLS)

        self._add_clinic_header(ws, "Patient Records", N)

        # Headers row 3
        for c, h in enumerate(COLS, 1):
            ws.cell(row=3, column=c, value=h)
        self._style_header_row(ws, 3, N, "2563EB")
        ws.row_dimensions[3].height = 22

        # Data
        def calc_age(dob):
            if not dob:
                return ""
            try:
                d = date.fromisoformat(str(dob)[:10])
                t = date.today()
                return t.year - d.year - ((t.month, t.day) < (d.month, d.day))
            except Exception:
                return ""

        for r, row in enumerate(rows, 4):
            vals = [
                row[0], row[1], str(row[2])[:10] if row[2] else "",
                calc_age(row[2]), row[3], row[7],
                row[4], row[5], row[6],
                row[9], row[8], str(row[10])[:10] if row[10] else "",
            ]
            for c, v in enumerate(vals, 1):
                ws.cell(row=r, column=c, value=v)

        self._style_data_rows(ws, 4, 3 + len(rows), N)

        # Column widths
        widths = [6, 28, 14, 6, 10, 10, 16, 26, 16, 30, 28, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Summary row
        sr = 4 + len(rows) + 1
        ws.cell(row=sr, column=1, value="Total Patients:")
        ws.cell(row=sr, column=1).font = Font(bold=True, name="Arial")
        ws.cell(row=sr, column=2, value=f"=COUNTA(B4:B{3+len(rows)})")
        ws.cell(row=sr, column=2).font = Font(bold=True, name="Arial", color="2563EB")

        ws.freeze_panes = "A4"
        wb.save(path)
        Toast.show(self._content, f"✅ Patients exported — {len(rows)} records", success=True)
        self._open_file(path)

    # ══════════════════════════════════════════════════════════════════════
    # EXPORT: APPOINTMENTS
    # ══════════════════════════════════════════════════════════════════════

    def _export_appointments(self):
        # Ask date range
        ApptExportDialog(self, self.controller, self._content)

    def _do_export_appointments(self, date_filter: str, label: str):
        path = self._ask_save(
            f"appointments_{label}_{datetime.now().strftime('%Y%m%d')}.xlsx")
        if not path:
            return

        try:
            rows = self.controller.db.fetch_all(
                f"""SELECT a.appointment_id, p.full_name,
                           a.appointment_date, a.status,
                           a.visit_type, a.chief_complaint,
                           a.diagnosis, a.total_amount
                    FROM appointments a
                    JOIN patients p ON a.patient_id = p.patient_id
                    {date_filter}
                    ORDER BY a.appointment_date DESC"""
            ) or []
        except Exception as e:
            messagebox.showerror("Export Error", str(e)); return


        wb = self._make_wb()
        ws = wb.active
        ws.title = "Appointments"

        COLS = ["#", "Patient", "Date", "Time", "Visit Type",
                "Status", "Chief Complaint", "Diagnosis", "Total (DA)"]
        N = len(COLS)

        self._add_clinic_header(ws, f"Appointments — {label}", N)

        for c, h in enumerate(COLS, 1):
            ws.cell(row=3, column=c, value=h)
        self._style_header_row(ws, 3, N, "7C3AED")
        ws.row_dimensions[3].height = 22

        STATUS_COLORS = {
            "Completed":   "DCFCE7",
            "Cancelled":   "FEE2E2",
            "Waiting":     "CCFBF1",
            "In Progress": "DBEAFE",
            "No Show":     "F3F4F6",
        }

        for r, row in enumerate(rows, 4):
            dt = str(row[2]) if row[2] else ""
            date_part = dt[:10]
            time_part = dt[11:16] if len(dt) >= 16 else ""
            vals = [row[0], row[1], date_part, time_part,
                    row[4], row[3], row[5], row[6], row[7]]
            for c, v in enumerate(vals, 1):
                ws.cell(row=r, column=c, value=v)

            # Colour status cell
            status = row[3]
            if status in STATUS_COLORS:
                fill = PatternFill("solid", fgColor=STATUS_COLORS[status])
                ws.cell(row=r, column=6).fill = fill

            # Right-align total
            ws.cell(row=r, column=9).alignment = Alignment(horizontal="right")
            ws.cell(row=r, column=9).number_format = '#,##0.00'

        self._style_data_rows(ws, 4, 3 + len(rows), N)

        widths = [6, 26, 12, 8, 16, 12, 28, 30, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Summary
        sr = 4 + len(rows) + 1
        ws.cell(row=sr, column=1, value="Total:").font = Font(bold=True, name="Arial")
        ws.cell(row=sr, column=9,
                value=f"=SUM(I4:I{3+len(rows)})"
                ).font = Font(bold=True, name="Arial", color="16A34A")
        ws.cell(row=sr, column=9).number_format = '#,##0.00'

        ws.freeze_panes = "A4"
        wb.save(path)
        Toast.show(self._content,
                   f"✅ Appointments exported — {len(rows)} records",
                   success=True)
        self._open_file(path)

    # ══════════════════════════════════════════════════════════════════════
    # EXPORT: PRESCRIPTIONS
    # ══════════════════════════════════════════════════════════════════════

    def _export_prescriptions(self):
        path = self._ask_save(
            f"prescriptions_{datetime.now().strftime('%Y%m%d')}.xlsx")
        if not path:
            return

        try:
            rows = self.controller.db.fetch_all(
                """SELECT r.rx_id, p.full_name, r.issued_date,
                          ri.medicine_name, ri.dosage, ri.frequency,
                          ri.duration, ri.instructions, r.notes
                   FROM prescriptions r
                   JOIN patients p ON r.patient_id = p.patient_id
                   LEFT JOIN prescription_items ri ON r.rx_id = ri.rx_id
                   ORDER BY r.issued_date DESC, r.rx_id, ri.item_id"""
            ) or []
        except Exception as e:
            messagebox.showerror("Export Error", str(e)); return


        wb = self._make_wb()
        ws = wb.active
        ws.title = "Prescriptions"

        COLS = ["Rx #", "Patient", "Date", "Medicine", "Dosage",
                "Frequency", "Duration", "Instructions", "Notes"]
        N = len(COLS)

        self._add_clinic_header(ws, "Prescriptions", N)
        for c, h in enumerate(COLS, 1):
            ws.cell(row=3, column=c, value=h)
        self._style_header_row(ws, 3, N, "16A34A")
        ws.row_dimensions[3].height = 22

        for r, row in enumerate(rows, 4):
            vals = [f"Rx-{row[0]:04d}", row[1],
                    str(row[2])[:10] if row[2] else "",
                    row[3], row[4], row[5], row[6], row[7], row[8]]
            for c, v in enumerate(vals, 1):
                ws.cell(row=r, column=c, value=v or "")

        self._style_data_rows(ws, 4, 3 + len(rows), N)

        widths = [10, 26, 12, 22, 12, 16, 12, 28, 24]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A4"
        wb.save(path)
        Toast.show(self._content,
                   f"✅ Prescriptions exported — {len(rows)} rows",
                   success=True)
        self._open_file(path)

    # ══════════════════════════════════════════════════════════════════════
    # EXPORT: REVENUE
    # ══════════════════════════════════════════════════════════════════════

    def _export_revenue(self):
        path = self._ask_save(
            f"revenue_{datetime.now().strftime('%Y%m')}.xlsx")
        if not path:
            return

        try:
            daily = self.controller.db.fetch_all(
                """SELECT DATE(appointment_date) AS day,
                          COUNT(*) AS total,
                          SUM(CASE WHEN status='Completed' THEN 1 ELSE 0 END) AS done,
                          COALESCE(SUM(CASE WHEN status='Completed'
                                      THEN total_amount ELSE 0 END),0) AS revenue
                   FROM appointments
                   GROUP BY day ORDER BY day DESC"""
            ) or []

            monthly = self.controller.db.fetch_all(
                """SELECT strftime('%Y-%m', appointment_date) AS month,
                          COUNT(*) AS total,
                          SUM(CASE WHEN status='Completed' THEN 1 ELSE 0 END) AS done,
                          COALESCE(SUM(CASE WHEN status='Completed'
                                      THEN total_amount ELSE 0 END),0) AS revenue
                   FROM appointments
                   GROUP BY month ORDER BY month DESC"""
            ) or []
        except Exception as e:
            messagebox.showerror("Export Error", str(e)); return


        wb = self._make_wb()

        # ── Sheet 1: Daily ────────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Daily Revenue"
        COLS1 = ["Date", "Total Appts", "Completed", "Revenue (DA)", "Avg per Visit (DA)"]
        N1 = len(COLS1)
        self._add_clinic_header(ws1, "Daily Revenue Report", N1)
        for c, h in enumerate(COLS1, 1):
            ws1.cell(row=3, column=c, value=h)
        self._style_header_row(ws1, 3, N1, "D97706")
        ws1.row_dimensions[3].height = 22

        for r, row in enumerate(daily, 4):
            revenue = float(row[3] or 0)
            done    = int(row[2] or 0)
            ws1.cell(row=r, column=1, value=str(row[0]))
            ws1.cell(row=r, column=2, value=int(row[1] or 0))
            ws1.cell(row=r, column=3, value=done)
            ws1.cell(row=r, column=4, value=revenue)
            ws1.cell(row=r, column=4).number_format = '#,##0.00'
            ws1.cell(row=r, column=5,
                     value=f"=IF(C{r}>0, D{r}/C{r}, 0)")
            ws1.cell(row=r, column=5).number_format = '#,##0.00'

        self._style_data_rows(ws1, 4, 3 + len(daily), N1)
        sr1 = 4 + len(daily) + 1
        for col, val in [(1,"TOTAL"), (2,f"=SUM(B4:B{3+len(daily)})"),
                         (3,f"=SUM(C4:C{3+len(daily)})"),
                         (4,f"=SUM(D4:D{3+len(daily)})")]:
            c = ws1.cell(row=sr1, column=col, value=val)
            c.font = Font(bold=True, name="Arial",
                           color="D97706" if col == 4 else "000000")
            if col in (2,3,4):
                c.number_format = '#,##0.00' if col == 4 else '0'

        for i, w in enumerate([14,14,12,16,18], 1):
            ws1.column_dimensions[get_column_letter(i)].width = w
        ws1.freeze_panes = "A4"

        # ── Sheet 2: Monthly ──────────────────────────────────────────────
        ws2 = wb.create_sheet("Monthly Revenue")
        COLS2 = ["Month", "Total Appts", "Completed", "Revenue (DA)", "Avg per Visit (DA)"]
        N2 = len(COLS2)
        self._add_clinic_header(ws2, "Monthly Revenue Summary", N2)
        for c, h in enumerate(COLS2, 1):
            ws2.cell(row=3, column=c, value=h)
        self._style_header_row(ws2, 3, N2, "16A34A")
        ws2.row_dimensions[3].height = 22

        for r, row in enumerate(monthly, 4):
            revenue = float(row[3] or 0)
            ws2.cell(row=r, column=1, value=str(row[0]))
            ws2.cell(row=r, column=2, value=int(row[1] or 0))
            ws2.cell(row=r, column=3, value=int(row[2] or 0))
            ws2.cell(row=r, column=4, value=revenue)
            ws2.cell(row=r, column=4).number_format = '#,##0.00'
            ws2.cell(row=r, column=5,
                     value=f"=IF(C{r}>0, D{r}/C{r}, 0)")
            ws2.cell(row=r, column=5).number_format = '#,##0.00'

        self._style_data_rows(ws2, 4, 3 + len(monthly), N2)
        for i, w in enumerate([14,14,12,16,18], 1):
            ws2.column_dimensions[get_column_letter(i)].width = w
        ws2.freeze_panes = "A4"

        wb.save(path)
        Toast.show(self._content, "✅ Revenue report exported", success=True)
        self._open_file(path)

    # ══════════════════════════════════════════════════════════════════════
    # EXPORT: SERVICES
    # ══════════════════════════════════════════════════════════════════════

    def _export_services(self):
        path = self._ask_save(
            f"services_{datetime.now().strftime('%Y%m%d')}.xlsx")
        if not path:
            return

        try:
            rows = self.controller.db.fetch_all(
                """SELECT code, name, category, price,
                          duration_min, description, is_active
                   FROM services ORDER BY category, name"""
            ) or []
        except Exception as e:
            messagebox.showerror("Export Error", str(e)); return


        wb = self._make_wb()
        ws = wb.active
        ws.title = "Services"

        COLS = ["Code", "Service Name", "Category",
                "Price (DA)", "Duration (min)", "Description", "Active"]
        N = len(COLS)

        self._add_clinic_header(ws, "Services Catalog", N)
        for c, h in enumerate(COLS, 1):
            ws.cell(row=3, column=c, value=h)
        self._style_header_row(ws, 3, N, "0D9488")
        ws.row_dimensions[3].height = 22

        for r, row in enumerate(rows, 4):
            vals = [row[0], row[1], row[2], float(row[3] or 0),
                    int(row[4] or 0), row[5],
                    "Yes" if row[6] else "No"]
            for c, v in enumerate(vals, 1):
                ws.cell(row=r, column=c, value=v)
            ws.cell(row=r, column=4).number_format = '#,##0.00'

        self._style_data_rows(ws, 4, 3 + len(rows), N)

        # Total row
        sr = 4 + len(rows) + 1
        ws.cell(row=sr, column=2, value="Total Services:").font = \
            Font(bold=True, name="Arial")
        ws.cell(row=sr, column=3,
                value=f"=COUNTA(C4:C{3+len(rows)})").font = \
            Font(bold=True, name="Arial", color="0D9488")

        widths = [10, 28, 16, 12, 14, 36, 8]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A4"
        wb.save(path)
        Toast.show(self._content,
                   f"✅ Services exported — {len(rows)} records",
                   success=True)
        self._open_file(path)

    # ══════════════════════════════════════════════════════════════════════
    # EXPORT: FULL REPORT (all sheets)
    # ══════════════════════════════════════════════════════════════════════

    def _export_full_report(self):
        if not self._check_openpyxl(): return
        path = self._ask_save(
            f"clinic_report_{datetime.now().strftime('%Y%m%d')}.xlsx")
        if not path:
            return

        # Temporarily redirect saves to the full workbook
        self._full_report_mode = True
        self._full_report_path = path

        from datetime        import date

        try:
            wb = Workbook()

            # ── Cover sheet ───────────────────────────────────────────────
            cover = wb.active
            cover.title = "Report Cover"
            try:
                clinic = self.controller.get_app_name()
            except Exception:
                clinic = "Clinic Manager"

            cover.merge_cells("A1:F1")
            c = cover["A1"]
            c.value     = clinic.upper()
            c.font      = Font(bold=True, size=20, name="Arial", color="FFFFFF")
            c.fill      = PatternFill("solid", fgColor="0D9488")
            c.alignment = Alignment(horizontal="center", vertical="center")
            cover.row_dimensions[1].height = 50

            cover.merge_cells("A2:F2")
            c2 = cover["A2"]
            c2.value     = f"Full Clinic Report — Generated {datetime.now().strftime('%d %B %Y, %H:%M')}"
            c2.font      = Font(size=11, name="Arial", italic=True, color="64748B")
            c2.fill      = PatternFill("solid", fgColor="F0FDF4")
            c2.alignment = Alignment(horizontal="center")
            cover.row_dimensions[2].height = 22

            # Quick stats
            stats = [
                ("Total Patients",     "SELECT COUNT(*) FROM patients"),
                ("Total Appointments", "SELECT COUNT(*) FROM appointments"),
                ("Total Prescriptions","SELECT COUNT(*) FROM prescriptions"),
                ("Total Revenue (DA)", "SELECT COALESCE(SUM(total_amount),0) FROM appointments WHERE status='Completed'"),
            ]
            cover.cell(row=4, column=1, value="Summary").font = \
                Font(bold=True, size=12, name="Arial")
            for i, (label, sql) in enumerate(stats, 5):
                try:
                    val = self.controller.db.fetch_one(sql, ())[0]
                except Exception:
                    val = "—"
                cover.cell(row=i, column=1, value=label).font = \
                    Font(name="Arial", size=10, color="64748B")
                cover.cell(row=i, column=2, value=val).font = \
                    Font(bold=True, name="Arial", size=11, color="0D9488")
                if i == 8:  # revenue
                    cover.cell(row=i, column=2).number_format = '#,##0.00'
            cover.column_dimensions["A"].width = 24
            cover.column_dimensions["B"].width = 18

            # ── Reuse individual export logic to fill sheets ──────────────
            # We'll use the db directly rather than re-run file dialogs

            def add_sheet_patients():
                ws = wb.create_sheet("Patients")
                rows = self.controller.db.fetch_all(
                    "SELECT patient_id,full_name,date_of_birth,gender,"
                    "phone,email,wilaya,blood_type,allergies,created_at "
                    "FROM patients ORDER BY full_name") or []
                COLS = ["ID","Full Name","DOB","Gender","Phone","Email",
                        "Wilaya","Blood","Allergies","Registered"]
                N = len(COLS)
                self._add_clinic_header(ws, "Patients", N)
                for c,h in enumerate(COLS,1): ws.cell(row=3,column=c,value=h)
                self._style_header_row(ws,3,N,"2563EB")
                for r,row in enumerate(rows,4):
                    for c,v in enumerate([row[0],row[1],str(row[2])[:10] if row[2] else "",
                                           row[3],row[4],row[5],row[6],row[7],row[8],
                                           str(row[9])[:10] if row[9] else ""],1):
                        ws.cell(row=r,column=c,value=v)
                self._style_data_rows(ws,4,3+len(rows),N)
                ws.freeze_panes="A4"
                return len(rows)

            def add_sheet_appointments():
                ws = wb.create_sheet("Appointments")
                rows = self.controller.db.fetch_all(
                    "SELECT a.appointment_id,p.full_name,a.appointment_date,"
                    "a.status,a.visit_type,a.chief_complaint,a.total_amount "
                    "FROM appointments a JOIN patients p ON a.patient_id=p.patient_id "
                    "ORDER BY a.appointment_date DESC") or []
                COLS=["#","Patient","Date","Time","Visit Type","Status","Complaint","Total DA"]
                N=len(COLS)
                self._add_clinic_header(ws,"Appointments",N)
                for c,h in enumerate(COLS,1): ws.cell(row=3,column=c,value=h)
                self._style_header_row(ws,3,N,"7C3AED")
                for r,row in enumerate(rows,4):
                    dt=str(row[2]) if row[2] else ""
                    for c,v in enumerate([row[0],row[1],dt[:10],dt[11:16] if len(dt)>=16 else "",
                                           row[4],row[3],row[5],float(row[6] or 0)],1):
                        ws.cell(row=r,column=c,value=v)
                    ws.cell(row=r,column=8).number_format='#,##0.00'
                self._style_data_rows(ws,4,3+len(rows),N)
                ws.freeze_panes="A4"
                return len(rows)

            def add_sheet_revenue():
                ws = wb.create_sheet("Revenue")
                rows = self.controller.db.fetch_all(
                    "SELECT strftime('%Y-%m',appointment_date) AS m,"
                    "COUNT(*),SUM(CASE WHEN status='Completed' THEN 1 ELSE 0 END),"
                    "COALESCE(SUM(CASE WHEN status='Completed' THEN total_amount ELSE 0 END),0) "
                    "FROM appointments GROUP BY m ORDER BY m DESC") or []
                COLS=["Month","Total Appts","Completed","Revenue (DA)","Avg per Visit"]
                N=len(COLS)
                self._add_clinic_header(ws,"Monthly Revenue",N)
                for c,h in enumerate(COLS,1): ws.cell(row=3,column=c,value=h)
                self._style_header_row(ws,3,N,"D97706")
                for r,row in enumerate(rows,4):
                    ws.cell(row=r,column=1,value=str(row[0]))
                    ws.cell(row=r,column=2,value=int(row[1] or 0))
                    ws.cell(row=r,column=3,value=int(row[2] or 0))
                    ws.cell(row=r,column=4,value=float(row[3] or 0))
                    ws.cell(row=r,column=4).number_format='#,##0.00'
                    ws.cell(row=r,column=5,value=f"=IF(C{r}>0,D{r}/C{r},0)")
                    ws.cell(row=r,column=5).number_format='#,##0.00'
                self._style_data_rows(ws,4,3+len(rows),N)
                ws.freeze_panes="A4"

            p_count = add_sheet_patients()
            a_count = add_sheet_appointments()
            add_sheet_revenue()

            wb.save(path)
            Toast.show(self._content,
                       f"✅ Full report exported — {p_count} patients, {a_count} appointments",
                       success=True)
            self._open_file(path)

        except Exception as e:
            messagebox.showerror("Export Error", str(e))


# ══════════════════════════════════════════════════════════════════════════════
# DIALOG — Appointment export date range picker
# ══════════════════════════════════════════════════════════════════════════════

class ApptExportDialog(ctk.CTkToplevel):

    def __init__(self, parent_view, controller, content_frame):
        super().__init__()
        self.parent_view   = parent_view
        self.controller    = controller
        self.content_frame = content_frame

        self.title("📅  Export Appointments")
        self.geometry("380x300")
        self.resizable(False, False)
        self.grab_set(); self.lift(); self.focus_force()

        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)
        self._build()

    def _build(self):
        fr = ctk.CTkFrame(self, fg_color=BG)
        fr.grid(row=0, column=0, sticky="nsew")
        fr.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(fr, text="Select date range to export:",
                     font=ctk.CTkFont(FONT,13,"bold"),
                     text_color=TEXT_PRIMARY
                     ).grid(row=0, column=0, padx=20, pady=(20,14), sticky="w")

        OPTIONS = [
            ("Today only",      "WHERE DATE(a.appointment_date)=DATE('now')", "today"),
            ("This week",       "WHERE a.appointment_date >= DATE('now','weekday 0','-7 days')", "this_week"),
            ("This month",      "WHERE strftime('%Y-%m',a.appointment_date)=strftime('%Y-%m','now')", "this_month"),
            ("Last 3 months",   "WHERE a.appointment_date >= DATE('now','-3 months')", "last_3m"),
            ("This year",       "WHERE strftime('%Y',a.appointment_date)=strftime('%Y','now')", "this_year"),
            ("All time",        "", "all"),
        ]
        self._choice = ctk.StringVar(value="all")
        self._options = OPTIONS

        for i, (label, _, key) in enumerate(OPTIONS):
            ctk.CTkRadioButton(fr, text=label,
                               variable=self._choice, value=key,
                               font=ctk.CTkFont(FONT,12),
                               text_color=TEXT_PRIMARY,
                               fg_color=ACCENT
                               ).grid(row=i+1, column=0,
                                      padx=24, pady=3, sticky="w")

        br = ctk.CTkFrame(fr, fg_color="transparent")
        br.grid(row=len(OPTIONS)+1, column=0, sticky="ew",
                padx=20, pady=(16,20))
        br.grid_columnconfigure(0, weight=1)
        br.grid_columnconfigure(1, weight=1)

        ctk.CTkButton(br, text="Export",
                      font=ctk.CTkFont(FONT,13,"bold"),
                      fg_color=ACCENT, hover_color=ACCENT_HOVER,
                      height=40, corner_radius=8,
                      command=self._export
                      ).grid(row=0, column=0, sticky="ew", padx=(0,6))

        ctk.CTkButton(br, text="Cancel",
                      font=ctk.CTkFont(FONT,13),
                      fg_color=PANEL_ALT, hover_color=BORDER,
                      text_color=TEXT_PRIMARY,
                      height=40, corner_radius=8,
                      command=self.destroy
                      ).grid(row=0, column=1, sticky="ew", padx=(6,0))

    def _export(self):
        key = self._choice.get()
        for label, sql_filter, k in self._options:
            if k == key:
                self.destroy()
                self.parent_view._do_export_appointments(sql_filter, k)
                return